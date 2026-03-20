#!/usr/bin/env python3
"""
Excel 任务执行器 - 读取 Excel 文件内容，让智能体在手机上执行任务

使用示例：
    python excel_task.py --file a.xlsx --task "打开微信，给峰峰峰回路转发送文档里的所有问题" --output result.xlsx
"""

import argparse
import base64
import json
import os
import sys
from pathlib import Path
from datetime import datetime

# Fix Windows console encoding
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Try to import pandas
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Setup simple logger
import logging
logger = logging.getLogger(__name__)

from phone_agent import PhoneAgent
from phone_agent.agent import AgentConfig
from phone_agent.config import get_system_prompt
from phone_agent.model import ModelConfig
from phone_agent.batch_runner import BatchQuestionRunner, BatchConfig
from phone_agent.device_factory import get_device_factory
from queue import Queue
import threading


def load_excel_content(file_path: str, column: str = None) -> str:
    """
    读取 Excel 文件内容

    Args:
        file_path: Excel 文件路径
        column: 指定读取的列名（可选）

    Returns:
        格式化的文本内容
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("需要安装 pandas: pip install pandas openpyxl")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")

    # 读取 Excel
    df = pd.read_excel(path)

    # 获取所有列名
    columns = df.columns.tolist()

    # 如果指定了列，只读取该列
    if column:
        if column not in columns:
            raise ValueError(f"列'{column}'不存在，可用列：{columns}")
        items = df[column].dropna().tolist()
        return "\n".join([f"{i+1}. {str(item)}" for i, item in enumerate(items)])

    # 否则读取所有列
    # 查找"问题"列（不区分大小写）
    question_col = None
    for col in columns:
        if '问题' in col.lower() or 'question' in col.lower():
            question_col = col
            break

    if question_col:
        items = df[question_col].dropna().tolist()
        return "\n".join([f"{i+1}. {str(item)}" for i, item in enumerate(items)])

    # 如果没有找到问题列，返回所有数据
    lines = []
    for idx, row in df.iterrows():
        row_str = "，".join([f"{col}: {row[col]}" for col in columns if pd.notna(row[col])])
        lines.append(f"{idx+1}. {row_str}")

    return "\n".join(lines)


def load_txt_content(file_path: str) -> str:
    """读取 TXT 文件内容（每行一个）"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")

    with open(path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    items = [line.strip() for line in lines if line.strip()]
    return "\n".join([f"{i+1}. {item}" for i, item in enumerate(items)])


def load_file_content(file_path: str, column: str = None) -> str:
    """根据文件类型读取内容"""
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in ['.xlsx', '.xls']:
        return load_excel_content(file_path, column)
    elif suffix == '.txt':
        return load_txt_content(file_path)
    else:
        raise ValueError(f"不支持的文件格式：{suffix}")


def load_config(config_path: str = "config.json") -> dict:
    """加载配置文件"""
    path = Path(config_path)
    if not path.exists():
        return {}

    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def process_excel_questions(
    excel_path: str,
    task_template: str,
    output_path: str,
    model_cfg: ModelConfig,
    agent_cfg: AgentConfig,
    save_screenshots: bool = False,
    screenshot_dir: str = "./excel_screenshots",
    embed_screenshot: bool = False,
    column: str = None,
    progress_callback: callable = None
) -> list:
    """
    逐条处理 Excel 中的问题

    Args:
        excel_path: Excel 文件路径
        task_template: 任务模板
        output_path: 输出 Excel 文件路径
        model_cfg: 模型配置
        agent_cfg: Agent 配置
        save_screenshots: 是否保存截图
        screenshot_dir: 截图保存目录
        embed_screenshot: 是否嵌入截图
        column: 指定读取的列名
        progress_callback: 进度回调函数

    Returns:
        执行结果列表
    """
    if not PANDAS_AVAILABLE:
        print("❌ 需要安装 pandas")
        return []

    # 读取 Excel 获取问题列表
    df = pd.read_excel(excel_path)

    # 确定问题列
    if column and column in df.columns:
        question_col = column
    else:
        # 自动查找"问题"列
        question_col = None
        for col in df.columns:
            if '问题' in col.lower() or 'question' in col.lower():
                question_col = col
                break
        if not question_col:
            question_col = df.columns[0]  # 默认第一列

    # 获取问题列表
    questions = df[question_col].dropna().astype(str).tolist()
    questions = [q.strip() for q in questions if q.strip() and q != 'nan']

    if not questions:
        print("❌ 没有找到任何问题")
        return []

    print(f"📋 共找到 {len(questions)} 个问题")

    # 创建 Agent 实例
    agent = PhoneAgent(model_config=model_cfg, agent_config=agent_cfg)

    results = []

    for i, question in enumerate(questions, 1):
        print(f"\n{'='*60}")
        print(f"问题 {i}/{len(questions)}: {question[:50]}...")
        print(f"{'='*60}")

        # 更新进度
        if progress_callback:
            progress_callback(i, len(questions), question)

        try:
            # 构建任务
            if "{content}" in task_template:
                full_task = task_template.replace("{content}", question)
            else:
                full_task = f"{task_template}\n\n问题：{question}"

            # 执行任务
            answer = agent.run(full_task)

            # 保存结果
            result = {
                'question': question,
                'answer': answer,
                'success': True,
                'screenshot_path': None,
                'steps': agent.step_count
            }

            # 保存截图（如果启用）
            if save_screenshots:
                try:
                    device_factory = get_device_factory()
                    screenshot = device_factory.get_screenshot(enable_compression=False)

                    # 保存截图文件
                    os.makedirs(screenshot_dir, exist_ok=True)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    screenshot_filename = f"screenshot_{timestamp}_{i}.png"
                    screenshot_path = os.path.join(screenshot_dir, screenshot_filename)

                    image_data = base64.b64decode(screenshot.base64_data)
                    with open(screenshot_path, "wb") as f:
                        f.write(image_data)

                    result['screenshot_path'] = screenshot_path
                    print(f"📸 截图已保存：{screenshot_path}")

                except Exception as e:
                    print(f"⚠️ 保存截图失败：{e}")

            results.append(result)
            print(f"✅ 完成：{answer[:50] if answer else '无结果'}...")

            # 重置 Agent 上下文
            agent.reset()

        except Exception as e:
            print(f"❌ 执行失败：{e}")
            results.append({
                'question': question,
                'answer': '',
                'success': False,
                'screenshot_path': None,
                'error': str(e)
            })

    # 保存结果到 Excel
    print(f"\n📊 保存结果到 {output_path}...")
    save_results_to_excel(excel_path, output_path, results, save_screenshots, embed_screenshot, screenshot_dir)

    return results


def save_results_to_excel(
    excel_path: str,
    output_path: str,
    results: list,
    save_screenshots: bool = False,
    embed_screenshot: bool = False
):
    """保存结果到 Excel 文件"""
    if not PANDAS_AVAILABLE:
        print("❌ pandas 不可用")
        return

    # 读取原始 Excel
    df = pd.read_excel(excel_path)

    # 添加或更新列
    if '答案' not in df.columns:
        df['答案'] = ''
    if '状态' not in df.columns:
        df['状态'] = ''
    if save_screenshots and '截图路径' not in df.columns:
        df['截图路径'] = ''

    # 更新数据
    for i, result in enumerate(results):
        if i < len(df):
            df.loc[i, '答案'] = result.get('answer', '')
            df.loc[i, '状态'] = '成功' if result.get('success', False) else f"失败：{result.get('error', '')}"
            if save_screenshots:
                df.loc[i, '截图路径'] = result.get('screenshot_path', '')

    # 保存
    df.to_excel(output_path, index=False, engine='openpyxl')
    print(f"✅ 结果已保存到：{output_path}")

    # 嵌入截图（如果启用）
    if embed_screenshot and save_screenshots:
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image
            from PIL import Image as PILImage

            wb = load_workbook(output_path)
            ws = wb.active

            # 找到截图路径列
            screenshot_col = None
            headers = {cell.column_letter: cell.value for cell in ws[1]}
            for col_letter, header in headers.items():
                if header == '截图路径':
                    screenshot_col = col_letter
                    break

            if screenshot_col:
                for i, result in enumerate(results):
                    screenshot_path = result.get('screenshot_path')
                    if screenshot_path and os.path.exists(screenshot_path):
                        row = i + 2  # Excel 行号从 1 开始，第 1 行是标题
                        try:
                            with PILImage.open(screenshot_path) as img:
                                max_width = 300
                                max_height = 400
                                width, height = img.size
                                ratio = min(max_width / width, max_height / height)
                                if ratio < 1:
                                    new_width = int(width * ratio * 0.75)
                                    new_height = int(height * ratio * 0.75)
                                else:
                                    new_width = int(width * 0.75)
                                    new_height = int(height * 0.75)

                            img_obj = Image(screenshot_path)
                            img_obj.width = new_width
                            img_obj.height = new_height
                            anchor = f"{screenshot_col}{row}"
                            img_obj.anchor = anchor
                            ws.add_image(img_obj)
                            ws.row_dimensions[row].height = new_height / 6
                        except Exception as e:
                            print(f"⚠️ 嵌入截图失败 ({screenshot_path}): {e}")

            wb.save(output_path)
            print(f"✅ 截图已嵌入到 Excel")
        except Exception as e:
            print(f"⚠️ 嵌入截图失败：{e}")


def save_result_to_excel(
    excel_path: str,
    output_path: str,
    answer: str,
    success: bool = True,
    error_message: str = None
) -> None:
    """
    仅保存任务结果到 Excel 文件（不保存截图）

    Args:
        excel_path: 原始 Excel 文件路径
        output_path: 输出 Excel 文件路径
        answer: 执行结果/答案
        success: 是否成功
        error_message: 错误消息
    """
    if not PANDAS_AVAILABLE:
        logger.warning("pandas not available, cannot save to Excel")
        return

    try:
        # 读取原始 Excel
        df = pd.read_excel(excel_path)

        # 检查是否需要添加列
        required_columns = ['答案', '状态']
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''

        # 更新第一行
        df.loc[0, '答案'] = answer
        df.loc[0, '状态'] = '成功' if success else f'失败：{error_message}'

        # 保存到输出文件
        df.to_excel(output_path, index=False, engine='openpyxl')

        print(f"✅ 结果已保存到：{output_path}")

    except Exception as e:
        print(f"❌ 保存结果到 Excel 失败：{e}")
        import traceback
        traceback.print_exc()


def save_screenshot_and_update_excel(
    excel_path: str,
    output_path: str,
    screenshot_dir: str,
    task_description: str,
    answer: str,
    success: bool = True,
    error_message: str = None,
    embed_screenshot: bool = True  # 是否将截图嵌入 Excel
) -> str:
    """
    保存当前屏幕截图，并将结果更新到 Excel 文件

    Args:
        excel_path: 原始 Excel 文件路径
        output_path: 输出 Excel 文件路径
        screenshot_dir: 截图保存目录
        task_description: 任务描述
        answer: 执行结果/答案
        success: 是否成功
        error_message: 错误消息
        embed_screenshot: 是否将截图嵌入 Excel

    Returns:
        截图文件路径
    """
    if not PANDAS_AVAILABLE:
        logger.warning("pandas not available, cannot save to Excel")
        return None

    try:
        # 保存截图
        os.makedirs(screenshot_dir, exist_ok=True)
        device_factory = get_device_factory()
        screenshot = device_factory.get_screenshot(enable_compression=False)

        # 生成截图文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        screenshot_filename = f"screenshot_{timestamp}_{abs(hash(task_description)) % 10000}.png"
        screenshot_path = os.path.join(screenshot_dir, screenshot_filename)

        # 从 base64_data 获取二进制数据
        image_data = base64.b64decode(screenshot.base64_data)

        with open(screenshot_path, "wb") as f:
            f.write(image_data)

        # 读取原始 Excel
        df = pd.read_excel(excel_path)

        # 检查是否需要添加列
        required_columns = ['答案', '截图路径', '状态']
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''

        # 更新第一行
        df.loc[0, '答案'] = answer
        df.loc[0, '截图路径'] = screenshot_path
        df.loc[0, '状态'] = '成功' if success else f'失败：{error_message}'

        # 保存到临时文件
        temp_output = output_path + '.tmp'
        df.to_excel(temp_output, index=False, engine='openpyxl')

        # 如果嵌入截图，使用 openpyxl 添加图片
        if embed_screenshot:
            embed_screenshot_to_excel(temp_output, output_path, screenshot_path, row=2)
            os.remove(temp_output)
        else:
            # 直接移动临时文件
            import shutil
            shutil.move(temp_output, output_path)

        print(f"✅ 截图已保存：{screenshot_path}")
        print(f"✅ 结果已更新到：{output_path}")
        if embed_screenshot:
            print(f"✅ 截图已嵌入到 Excel 文档中")

        return screenshot_path

    except Exception as e:
        print(f"❌ 保存截图或更新 Excel 失败：{e}")
        return None


def embed_screenshot_to_excel(excel_path: str, output_path: str, image_path: str, row: int = 2):
    """
    将截图嵌入到 Excel 文件的指定行

    Args:
        excel_path: 输入的 Excel 文件路径
        output_path: 输出的 Excel 文件路径
        image_path: 截图文件路径
        row: 插入图片的行号（从 1 开始，默认为 2 即数据第一行）
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    from PIL import Image as PILImage

    wb = load_workbook(excel_path)
    ws = wb.active

    # 获取截图列（"截图" 列）
    screenshot_col = 'D'  # 默认 D 列是截图路径

    # 获取列名
    headers = {cell.column_letter: cell.value for cell in ws[1]}
    for col_letter, header in headers.items():
        if header == '截图路径' or header == '截图':
            screenshot_col = col_letter
            break

    # 打开图片获取尺寸
    with PILImage.open(image_path) as img:
        # 压缩图片尺寸以适应单元格
        max_width = 400
        max_height = 600
        width, height = img.size

        # 计算缩放比例
        ratio = min(max_width / width, max_height / height)
        if ratio < 1:
            new_width = int(width * ratio * 0.75)
            new_height = int(height * ratio * 0.75)
        else:
            new_width = int(width * 0.75)
            new_height = int(height * 0.75)

    # 创建图片对象
    img = Image(image_path)
    img.width = new_width
    img.height = new_height

    # 计算单元格位置
    anchor = f"{screenshot_col}{row}"
    img.anchor = anchor
    ws.add_image(img)

    # 调整行高
    ws.row_dimensions[row].height = new_height / 6

    # 保存
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="读取 Excel/TXT 文件内容，让智能体在手机上执行任务",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    # 读取 Excel 问题列表，打开微信发送给某人
    python excel_task.py --file 工作簿 2.xlsx --task "打开微信，给峰峰峰回路转发送文档里的所有问题"

    # 指定读取的列
    python excel_task.py --file questions.xlsx --column 问题 --task "打开微信，把所有问题发给张三"

    # 读取 TXT 文件
    python excel_task.py --file questions.txt --task "在抖音搜索文档里的每个问题"

    # 逐条执行 Excel 中的每个问题（推荐，不会遗漏）
    python excel_task.py --file questions.xlsx --task "请回答这个问题：{content}" --mode batch

    # 批量执行并保存截图
    python excel_task.py --file questions.xlsx --task "请回答：{content}" --mode batch --save-screenshots --embed-screenshot
        """
    )

    parser.add_argument(
        "--file", "-f",
        type=str,
        required=True,
        help="Excel 或 TXT 文件路径"
    )

    parser.add_argument(
        "--task", "-t",
        type=str,
        required=True,
        help="任务描述，可以使用 {content} 占位符引用文件内容"
    )

    parser.add_argument(
        "--column", "-c",
        type=str,
        default=None,
        help="指定读取的 Excel 列名（默认自动检测'问题'列）"
    )

    parser.add_argument(
        "--preview",
        action="store_true",
        help="只预览文件内容，不执行任务"
    )

    parser.add_argument(
        "--config",
        type=str,
        default="config.json",
        help="配置文件路径（默认：config.json）"
    )

    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="输出 Excel 文件路径（默认：在原文件名后添加 _results）"
    )

    parser.add_argument(
        "--save-screenshots",
        action="store_true",
        default=False,
        help="保存截图到 Excel（默认不保存，只保存答案文本）"
    )

    parser.add_argument(
        "--no-screenshots",
        action="store_true",
        default=False,
        help="不保存截图（默认行为，此参数用于显式禁用截图）"
    )

    parser.add_argument(
        "--screenshot-dir",
        type=str,
        default="./excel_screenshots",
        help="截图保存目录（默认：./excel_screenshots）"
    )

    parser.add_argument(
        "--embed-screenshot",
        action="store_true",
        default=False,
        help="将截图直接嵌入到 Excel 文档中（而不只是保存路径）"
    )

    parser.add_argument(
        "--mode",
        type=str,
        choices=["single", "batch"],
        default="single",
        help="执行模式：single=单个任务（默认），batch=逐条执行 Excel 中的每个问题"
    )

    parser.add_argument(
        "--max-questions",
        type=int,
        default=0,
        help="最大执行问题数（0=全部，仅在 batch 模式下有效）"
    )

    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="显示详细输出"
    )

    args = parser.parse_args()

    # 检查 pandas
    if not PANDAS_AVAILABLE and args.file.lower().endswith(('.xlsx', '.xls')):
        print("❌ 错误：读取 Excel 需要安装 pandas")
        print("   请运行：pip install pandas openpyxl")
        sys.exit(1)

    # 加载配置
    config = load_config(args.config)
    model_config_dict = config.get('model', {})
    agent_config_dict = config.get('agent', {})

    # 创建模型配置
    model_cfg = ModelConfig(
        base_url=model_config_dict.get('base_url', 'http://localhost:11434/v1'),
        model_name=model_config_dict.get('model_name', 'qwen3.5:4b'),
        api_key=model_config_dict.get('api_key', 'ollama'),
        use_thinking=model_config_dict.get('use_thinking', False),
        lang=agent_config_dict.get('lang', 'cn'),
    )

    # 创建 Agent 配置
    agent_cfg = AgentConfig(
        max_steps=agent_config_dict.get('max_steps', 50),
        verbose=args.verbose,
        lang=agent_config_dict.get('lang', 'cn'),
    )

    # 确定输出文件路径
    if args.output:
        output_file = args.output
    else:
        # 默认在原文件名后添加 _results
        input_path = Path(args.file)
        output_file = str(input_path.parent / f"{input_path.stem}_results{input_path.suffix}")

    # 预览模式
    if args.preview:
        print(f"📄 读取文件：{args.file}")
        try:
            content = load_file_content(args.file, args.column)
        except Exception as e:
            print(f"❌ 读取文件失败：{e}")
            sys.exit(1)
        print("\n📋 文件内容预览：")
        print("-" * 50)
        lines = content.split("\n")
        for line in lines[:10]:
            print(f"  {line}")
        if len(lines) > 10:
            print(f"  ... 还有 {len(lines) - 10} 行")
        print("-" * 50)
        return

    # 根据模式执行
    if args.mode == "batch":
        # 批量执行模式 - 逐条执行 Excel 中的问题
        run_batch_mode(args, model_cfg, agent_cfg, output_file)
    else:
        # 单个任务模式 - 将所有问题合并成一个任务
        run_single_mode(args, model_cfg, agent_cfg, output_file)


def run_single_mode(args, model_cfg, agent_cfg, output_file):
    """单个任务模式：将所有问题合并成一个任务执行"""
    # 加载文件内容
    print(f"📄 读取文件：{args.file}")
    try:
        content = load_file_content(args.file, args.column)
    except Exception as e:
        print(f"❌ 读取文件失败：{e}")
        sys.exit(1)

    # 构建任务提示
    if "{content}" in args.task:
        full_task = f"""【重要说明】
我已经从电脑上的文件读取了以下内容，请直接使用这些内容执行任务。
注意：文件已经读取完成，不需要在手机上查找或打开任何文件。

文件内容：
---
{content}
---

用户任务：{args.task.replace("{content}", "上述内容")}
"""
    else:
        full_task = f"""【重要说明】
我已经从电脑上的文件 "{args.file}" 中读取了所有内容。
注意：文件已经读取完成，不需要在手机上查找或打开任何文件。
请直接根据以下文件内容执行任务。

=== 文件内容开始 ===
{content}
=== 文件内容结束 ===

用户任务：{args.task}

请按照上述任务要求，在手机执行相应操作。
如果需要发送文件内容，请直接逐条发送上述列表中的内容。
"""

    print("\n🤖 开始执行任务...")
    print(f"   任务：{args.task[:50]}..." if len(args.task) > 50 else f"   任务：{args.task}")
    print(f"   文件内容：共 {len(content.split(chr(10)))} 行")
    print(f"   输出文件：{output_file}")
    if args.save_screenshots:
        print(f"   保存截图：是")
        print(f"   截图目录：{args.screenshot_dir}")
        if args.embed_screenshot:
            print(f"   嵌入截图：是（截图将直接显示在 Excel 中）")
    else:
        print(f"   保存截图：否（只保存答案文本）")
    print("-" * 50)

    # 创建并运行 Agent
    agent = PhoneAgent(model_config=model_cfg, agent_config=agent_cfg)

    try:
        result = agent.run(full_task)

        if args.save_screenshots:
            print(f"\n📸 保存截图...")
            screenshot_path = save_screenshot_and_update_excel(
                excel_path=args.file,
                output_path=output_file,
                screenshot_dir=args.screenshot_dir,
                task_description=args.task,
                answer=result,
                success=True,
                embed_screenshot=args.embed_screenshot
            )
        else:
            print(f"\n💾 保存结果...")
            save_result_to_excel(
                excel_path=args.file,
                output_path=output_file,
                answer=result,
                success=True
            )
            screenshot_path = None

        print("\n" + "=" * 50)
        print("✅ 任务执行完成")
        print("=" * 50)
        print(f"输出文件：{output_file}")
        if screenshot_path:
            print(f"截图路径：{screenshot_path}")
            if args.embed_screenshot:
                print(f"✅ 截图已嵌入到 Excel 文档中，打开 {output_file} 即可看到截图")
        print("=" * 50)

    except KeyboardInterrupt:
        print("\n\n⚠️  用户中断")
    except Exception as e:
        print(f"\n❌ 任务失败：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def run_batch_mode(args, model_cfg, agent_cfg, output_file):
    """批量执行模式：逐条执行 Excel 中的每个问题"""
    print("\n" + "=" * 60)
    print("📊 批量执行模式 - 逐条处理 Excel 中的问题")
    print("=" * 60)
    print(f"   输入文件：{args.file}")
    print(f"   输出文件：{output_file}")
    print(f"   任务模板：{args.task[:50]}..." if len(args.task) > 50 else f"   任务模板：{args.task}")
    print(f"   保存截图：{'是' if args.save_screenshots else '否'}")
    if args.save_screenshots:
        print(f"   截图目录：{args.screenshot_dir}")
        print(f"   嵌入截图：{'是' if args.embed_screenshot else '否'}")
    if args.max_questions > 0:
        print(f"   最大问题数：{args.max_questions}")
    print("=" * 60)

    try:
        # 执行批量处理
        results = process_excel_questions(
            excel_path=args.file,
            task_template=args.task,
            output_path=output_file,
            model_cfg=model_cfg,
            agent_cfg=agent_cfg,
            save_screenshots=args.save_screenshots,
            screenshot_dir=args.screenshot_dir,
            embed_screenshot=args.embed_screenshot,
            column=args.column
        )

        # 打印统计
        success_count = sum(1 for r in results if r.get('success', False))
        failed_count = len(results) - success_count

        print("\n" + "=" * 60)
        print("✅ 批量执行完成")
        print("=" * 60)
        print(f"总问题数：{len(results)}")
        print(f"成功：{success_count}")
        print(f"失败：{failed_count}")
        print(f"成功率：{success_count/len(results)*100:.1f}%" if results else "N/A")
        print(f"输出文件：{output_file}")
        print("=" * 60)

    except KeyboardInterrupt:
        print("\n\n⚠️  用户中断")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 批量执行失败：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
