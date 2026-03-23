#!/usr/bin/env python3
"""
Excel 任务执行器 - 读取 Excel 文件内容，让智能体在手机上执行任务

使用示例：
    # 批量执行并嵌入截图
    python bin/excel_task.py --file questions.xlsx --task "请回答这个问题：{content}" --mode batch --embed-screenshot

    # 只保存答案文本（默认）
    python bin/excel_task.py --file questions.xlsx --task "请回答这个问题：{content}" --mode batch

    # 对比 AI 回复和标准答案
    python bin/excel_task.py --file questions.xlsx --task "请回答这个问题：{content}" --mode batch --compare-answer
"""

import argparse
import base64
import json
import os
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

# Fix Windows console encoding
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Try to import pandas
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

import logging
logger = logging.getLogger(__name__)

from phone_agent import PhoneAgent
from phone_agent.agent import AgentConfig
from phone_agent.model import ModelConfig
from phone_agent.device_factory import get_device_factory


def load_excel_content(file_path: str, column: str = None) -> str:
    """读取 Excel 文件内容"""
    if not PANDAS_AVAILABLE:
        raise ImportError("需要安装 pandas: pip install pandas openpyxl")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")

    df = pd.read_excel(path)
    columns = df.columns.tolist()

    if column:
        if column not in columns:
            raise ValueError(f"列'{column}'不存在，可用列：{columns}")
        items = df[column].dropna().tolist()
        return "\n".join([f"{i+1}. {str(item)}" for i, item in enumerate(items)])

    # 自动查找"问题"列
    question_col = None
    for col in columns:
        if '问题' in col.lower() or 'question' in col.lower():
            question_col = col
            break

    if question_col:
        items = df[question_col].dropna().tolist()
        return "\n".join([f"{i+1}. {str(item)}" for i, item in enumerate(items)])

    # 返回所有数据
    lines = []
    for idx, row in df.iterrows():
        row_str = "，".join([f"{col}: {row[col]}" for col in columns if pd.notna(row[col])])
        lines.append(f"{idx+1}. {row_str}")

    return "\n".join(lines)


def load_txt_content(file_path: str) -> str:
    """读取 TXT 文件内容（每行一个）"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")

    with open(path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    items = [line.strip() for line in lines if line.strip()]
    return "\n".join([f"{i+1}. {item}" for i, item in enumerate(items)])


def load_file_content(file_path: str, column: str = None) -> str:
    """根据文件类型读取内容"""
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in ['.xlsx', '.xls']:
        return load_excel_content(file_path, column)
    elif suffix == '.txt':
        return load_txt_content(file_path)
    else:
        raise ValueError(f"不支持的文件格式：{suffix}")


def load_config(config_path: str = "config.json") -> dict:
    """加载配置文件"""
    path = Path(config_path)
    if not path.exists():
        return {}

    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def process_excel_questions(
    excel_path: str,
    task_template: str,
    output_path: str,
    model_cfg: ModelConfig,
    agent_cfg: AgentConfig,
    embed_screenshot: bool = False,
    compare_answer: bool = False,
    column: str = None,
    progress_callback: callable = None
) -> list:
    """
    逐条处理 Excel 中的问题

    Args:
        excel_path: Excel 文件路径
        task_template: 任务模板
        output_path: 输出 Excel 文件路径
        model_cfg: 模型配置
        agent_cfg: Agent 配置
        embed_screenshot: 是否嵌入截图到 Excel
        compare_answer: 是否对比 AI 回复和标准答案（由智能体自己对比）
        column: 指定读取的列名
        progress_callback: 进度回调函数

    Returns:
        执行结果列表
    """
    if not PANDAS_AVAILABLE:
        print("❌ 需要安装 pandas")
        return []

    df = pd.read_excel(excel_path)

    # 确定问题列
    if column and column in df.columns:
        question_col = column
    else:
        question_col = None
        for col in df.columns:
            if '问题' in col.lower() or 'question' in col.lower():
                question_col = col
                break
        if not question_col:
            question_col = df.columns[0]

    # 查找标准答案列
    standard_answer_col = None
    if compare_answer:
        for col in df.columns:
            if '标准答案' in col.lower() or 'standard' in col.lower():
                standard_answer_col = col
                break
        if not standard_answer_col:
            print("⚠️ 未找到'标准答案'列，将不进行答案对比")
            compare_answer = False

    questions = df[question_col].dropna().astype(str).tolist()
    questions = [q.strip() for q in questions if q.strip() and q != 'nan']

    # 获取标准答案
    standard_answers = []
    if compare_answer:
        standard_answers = df[standard_answer_col].dropna().astype(str).tolist()
        standard_answers = [s.strip() for s in standard_answers if s.strip() and s != 'nan']

    if not questions:
        print("❌ 没有找到任何问题")
        return []

    print(f"📋 共找到 {len(questions)} 个问题")
    if compare_answer:
        print(f"📊 已加载 {len(standard_answers)} 个标准答案")

    agent = PhoneAgent(model_config=model_cfg, agent_config=agent_cfg)
    results = []

    for i, question in enumerate(questions, 1):
        print(f"\n{'='*60}")
        print(f"问题 {i}/{len(questions)}: {question[:50]}...")
        print(f"{'='*60}")

        if progress_callback:
            progress_callback(i, len(questions), question)

        try:
            # 构建任务，包含标准答案供智能体对比
            if compare_answer and i-1 < len(standard_answers):
                standard_answer = standard_answers[i-1]
                if "{content}" in task_template:
                    full_task = f"""{task_template.replace("{content}", question)}

【答案对比任务】
标准答案：{standard_answer}

请你对比上面获取的答案和标准答案，计算相似度（0-100 的分数），考虑：
1. 关键信息是否一致
2. 核心要点是否覆盖

请输出 JSON 格式的结果：
{{"答案": "你的答案", "相似度": 85}}
"""
                else:
                    full_task = f"""{task_template}

问题：{question}

【答案对比任务】
标准答案：{standard_answer}

请你对比上面获取的答案和标准答案，计算相似度（0-100 的分数），考虑：
1. 关键信息是否一致
2. 核心要点是否覆盖

请输出 JSON 格式的结果：
{{"答案": "你的答案", "相似度": 85}}
"""
            else:
                if "{content}" in task_template:
                    full_task = task_template.replace("{content}", question)
                else:
                    full_task = f"{task_template}\n\n问题：{question}"

            answer = agent.run(full_task)

            # 尝试从答案中提取相似度
            similarity_result = None
            if compare_answer:
                try:
                    import re
                    json_match = re.search(r'\{.*"相似度"\s*:\s*(\d+).*\}', answer, re.DOTALL)
                    if json_match:
                        similarity_result = {
                            'overall_similarity': int(json_match.group(1)),
                            'combined_similarity': int(json_match.group(1))
                        }
                        print(f"🔍 智能体评估的相似度：{similarity_result['overall_similarity']}%")
                except Exception as e:
                    print(f"⚠️ 提取相似度失败：{e}")

            result = {
                'question': question,
                'answer': answer,
                'standard_answer': standard_answers[i-1] if compare_answer and i-1 < len(standard_answers) else None,
                'similarity': similarity_result,
                'success': True,
                'screenshot_base64': None,
                'steps': agent.step_count
            }

            # 获取截图（如果启用嵌入）
            if embed_screenshot:
                try:
                    device_factory = get_device_factory()
                    screenshot = device_factory.get_screenshot(enable_compression=False)
                    result['screenshot_base64'] = screenshot.base64_data
                    print(f"📸 已获取截图")
                except Exception as e:
                    print(f"⚠️ 获取截图失败：{e}")

            results.append(result)
            print(f"✅ 完成：{answer[:50] if answer else '无结果'}...")

            agent.reset()

        except Exception as e:
            print(f"❌ 执行失败：{e}")
            results.append({
                'question': question,
                'answer': '',
                'standard_answer': None,
                'similarity': None,
                'success': False,
                'screenshot_base64': None,
                'error': str(e)
            })

    # 保存结果到 Excel
    print(f"\n📊 保存结果到 {output_path}...")
    save_results_to_excel(excel_path, output_path, results, embed_screenshot, compare_answer)

    return results


def save_results_to_excel(
    excel_path: str,
    output_path: str,
    results: list,
    embed_screenshot: bool = False,
    compare_answer: bool = False
):
    """保存结果到 Excel 文件，支持嵌入截图和答案对比"""
    if not PANDAS_AVAILABLE:
        print("❌ pandas 不可用")
        return

    df = pd.read_excel(excel_path)

    # 添加或更新列
    if '答案' not in df.columns:
        df['答案'] = ''
    if '状态' not in df.columns:
        df['状态'] = ''

    # 答案对比相关列
    if compare_answer:
        if '相似度' not in df.columns:
            df['相似度'] = ''

    if embed_screenshot and '截图' not in df.columns:
        df['截图'] = ''

    # 更新数据
    for i, result in enumerate(results):
        if i < len(df):
            df.loc[i, '答案'] = result.get('answer', '')
            df.loc[i, '状态'] = '成功' if result.get('success', False) else f"失败：{result.get('error', '')}"

            # 保存相似度结果
            if compare_answer and result.get('similarity'):
                df.loc[i, '相似度'] = result['similarity'].get('overall_similarity', '')

    # 保存 Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
    print(f"✅ 结果已保存到：{output_path}")

    # 保存 Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
    print(f"✅ 结果已保存到：{output_path}")

    # 嵌入截图到单独的"截图"列
    if embed_screenshot:
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image
            from PIL import Image as PILImage
            import io

            wb = load_workbook(output_path)
            ws = wb.active

            # 找到截图列（单独的列）
            screenshot_col = None
            headers = {cell.column_letter: cell.value for cell in ws[1]}
            for col_letter, header in headers.items():
                if header == '截图':
                    screenshot_col = col_letter
                    break

            temp_paths = []  # 保存临时文件路径

            if screenshot_col:
                for i, result in enumerate(results):
                    screenshot_b64 = result.get('screenshot_base64')
                    if screenshot_b64:
                        row = i + 2
                        try:
                            # 解码 base64 图片
                            image_data = base64.b64decode(screenshot_b64)
                            img = PILImage.open(io.BytesIO(image_data))

                            # 缩放图片
                            max_width = 300
                            max_height = 400
                            width, height = img.size
                            ratio = min(max_width / width, max_height / height)
                            if ratio < 1:
                                new_width = int(width * ratio * 0.75)
                                new_height = int(height * ratio * 0.75)
                            else:
                                new_width = int(width * 0.75)
                                new_height = int(height * 0.75)

                            # 保存图片到临时文件
                            temp_path = Path(output_path).parent / f"temp_screenshot_{i}.png"
                            img.save(temp_path)
                            temp_paths.append(temp_path)

                            # 嵌入到 Excel 的截图列
                            img_obj = Image(temp_path)
                            img_obj.width = new_width
                            img_obj.height = new_height
                            anchor = f"{screenshot_col}{row}"
                            img_obj.anchor = anchor
                            ws.add_image(img_obj)
                            ws.row_dimensions[row].height = new_height / 6

                        except Exception as e:
                            print(f"⚠️ 嵌入截图失败 (行{row}): {e}")

            # 先保存 Excel
            wb.save(output_path)

            # 保存后再删除临时文件
            for temp_path in temp_paths:
                try:
                    if temp_path.exists():
                        os.remove(temp_path)
                except:
                    pass

            print(f"✅ 截图已嵌入到 Excel")
        except Exception as e:
            print(f"⚠️ 嵌入截图失败：{e}")


def save_result_to_excel(
    excel_path: str,
    output_path: str,
    answer: str,
    success: bool = True,
    error_message: str = None
) -> None:
    """保存单个任务结果到 Excel"""
    if not PANDAS_AVAILABLE:
        logger.warning("pandas not available, cannot save to Excel")
        return

    try:
        df = pd.read_excel(excel_path)

        required_columns = ['答案', '状态']
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''

        df.loc[0, '答案'] = answer
        df.loc[0, '状态'] = '成功' if success else f'失败：{error_message}'

        df.to_excel(output_path, index=False, engine='openpyxl')
        print(f"✅ 结果已保存到：{output_path}")

    except Exception as e:
        print(f"❌ 保存结果到 Excel 失败：{e}")
        import traceback
        traceback.print_exc()


def save_screenshot_and_update_excel(
    excel_path: str,
    output_path: str,
    answer: str,
    success: bool = True,
    error_message: str = None,
    embed_screenshot: bool = True
) -> str:
    """保存截图并更新 Excel（截图直接嵌入，不保存路径）"""
    import io as io_module

    if not PANDAS_AVAILABLE:
        logger.warning("pandas not available, cannot save to Excel")
        return None

    try:
        # 获取截图
        device_factory = get_device_factory()
        screenshot = device_factory.get_screenshot(enable_compression=False)
        screenshot_b64 = screenshot.base64_data

        # 读取 Excel
        df = pd.read_excel(excel_path)

        required_columns = ['答案', '状态']
        if embed_screenshot:
            required_columns.append('截图')
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''

        df.loc[0, '答案'] = answer
        df.loc[0, '状态'] = '成功' if success else f'失败：{error_message}'

        # 保存临时文件
        temp_output = output_path + '.tmp'
        df.to_excel(temp_output, index=False, engine='openpyxl')

        # 嵌入截图到单独的"截图"列
        if embed_screenshot:
            try:
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image
                from PIL import Image as PILImage

                wb = load_workbook(temp_output)
                ws = wb.active

                # 找到截图列
                screenshot_col = 'D'
                headers = {cell.column_letter: cell.value for cell in ws[1]}
                for col_letter, header in headers.items():
                    if header == '截图':
                        screenshot_col = col_letter
                        break

                # 解码并保存图片
                image_data = base64.b64decode(screenshot_b64)
                img = PILImage.open(io.BytesIO(image_data))

                max_width = 400
                max_height = 600
                width, height = img.size
                ratio = min(max_width / width, max_height / height)
                if ratio < 1:
                    new_width = int(width * ratio * 0.75)
                    new_height = int(height * ratio * 0.75)
                else:
                    new_width = int(width * 0.75)
                    new_height = int(height * 0.75)

                temp_img_path = Path(output_path).parent / "temp_screenshot.png"
                img.save(temp_img_path)

                img_obj = Image(temp_img_path)
                img_obj.width = new_width
                img_obj.height = new_height
                anchor = f"{screenshot_col}2"
                img_obj.anchor = anchor
                ws.add_image(img_obj)
                ws.row_dimensions[2].height = new_height / 6

                # 先保存再删除临时图片
                wb.save(output_path)
                os.remove(temp_img_path)
                print(f"✅ 截图已嵌入到 Excel")

            except Exception as e:
                print(f"⚠️ 嵌入截图失败：{e}")
                import shutil
                shutil.move(temp_output, output_path)
        else:
            import shutil
            shutil.move(temp_output, output_path)

        print(f"✅ 结果已更新到：{output_path}")
        return None

    except Exception as e:
        print(f"❌ 保存截图或更新 Excel 失败：{e}")
        return None


def main():
    parser = argparse.ArgumentParser(
        description="读取 Excel/TXT 文件内容，让智能体在手机上执行任务",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    # 批量执行 Excel 中的问题（只保存答案）
    python excel_task.py --file questions.xlsx --task "请回答这个问题：{content}" --mode batch

    # 批量执行并嵌入截图
    python excel_task.py --file questions.xlsx --task "请回答：{content}" --mode batch --embed-screenshot

    # 对比 AI 回复和标准答案（计算相似度）
    python excel_task.py --file questions.xlsx --task "请回答这个问题：{content}" --mode batch --compare-answer

    # 单个任务模式
    python excel_task.py --file questions.xlsx --task "打开微信，把所有问题发给张三"
        """
    )

    parser.add_argument("--file", "-f", type=str, required=True, help="Excel 或 TXT 文件路径")
    parser.add_argument("--task", "-t", type=str, required=True, help="任务描述，可使用 {content} 占位符")
    parser.add_argument("--column", "-c", type=str, default=None, help="指定读取的 Excel 列名")
    parser.add_argument("--preview", action="store_true", help="只预览文件内容，不执行任务")
    parser.add_argument("--config", type=str, default="config.json", help="配置文件路径")
    parser.add_argument("--output", "-o", type=str, default=None, help="输出 Excel 文件路径")
    parser.add_argument("--embed-screenshot", action="store_true", default=False, help="将截图嵌入到 Excel 文档中")
    parser.add_argument("--compare-answer", action="store_true", default=False, help="对比 AI 回复和标准答案并计算相似度")
    parser.add_argument("--mode", type=str, choices=["single", "batch"], default="single", help="执行模式")
    parser.add_argument("--max-questions", type=int, default=0, help="最大问题数（0=全部，batch 模式有效）")
    parser.add_argument("--verbose", "-v", action="store_true", help="显示详细输出")

    args = parser.parse_args()

    if not PANDAS_AVAILABLE and args.file.lower().endswith(('.xlsx', '.xls')):
        print("❌ 错误：读取 Excel 需要安装 pandas")
        print("   请运行：pip install pandas openpyxl")
        sys.exit(1)

    config = load_config(args.config)
    model_config_dict = config.get('model', {})
    agent_config_dict = config.get('agent', {})

    model_cfg = ModelConfig(
        base_url=model_config_dict.get('base_url', 'http://localhost:11434/v1'),
        model_name=model_config_dict.get('model_name', 'qwen3.5:4b'),
        api_key=model_config_dict.get('api_key', 'ollama'),
        use_thinking=model_config_dict.get('use_thinking', False),
        lang=agent_config_dict.get('lang', 'cn'),
    )

    agent_cfg = AgentConfig(
        max_steps=agent_config_dict.get('max_steps', 50),
        verbose=args.verbose,
        lang=agent_config_dict.get('lang', 'cn'),
    )

    # 确定输出文件
    if args.output:
        output_file = args.output
    else:
        input_path = Path(args.file)
        output_file = str(input_path.parent / f"{input_path.stem}_results{input_path.suffix}")

    # 预览模式
    if args.preview:
        print(f"📄 读取文件：{args.file}")
        try:
            content = load_file_content(args.file, args.column)
        except Exception as e:
            print(f"❌ 读取文件失败：{e}")
            sys.exit(1)
        print("\n📋 文件内容预览：")
        print("-" * 50)
        lines = content.split("\n")
        for line in lines[:10]:
            print(f"  {line}")
        if len(lines) > 10:
            print(f"  ... 还有 {len(lines) - 10} 行")
        print("-" * 50)
        return

    # 执行模式
    if args.mode == "batch":
        run_batch_mode(args, model_cfg, agent_cfg, output_file)
    else:
        run_single_mode(args, model_cfg, agent_cfg, output_file)


def run_single_mode(args, model_cfg, agent_cfg, output_file):
    """单个任务模式"""
    print(f"📄 读取文件：{args.file}")
    try:
        content = load_file_content(args.file, args.column)
    except Exception as e:
        print(f"❌ 读取文件失败：{e}")
        sys.exit(1)

    if "{content}" in args.task:
        full_task = f"""【重要说明】
我已经从电脑上的文件读取了以下内容，请直接使用这些内容执行任务。
注意：文件已经读取完成，不需要在手机上查找或打开任何文件。

文件内容：
---
{content}
---

用户任务：{args.task.replace("{content}", "上述内容")}
"""
    else:
        full_task = f"""【重要说明】
我已经从电脑上的文件 "{args.file}" 中读取了所有内容。
注意：文件已经读取完成，不需要在手机上查找或打开任何文件。
请直接根据以下文件内容执行任务。

=== 文件内容开始 ===
{content}
=== 文件内容结束 ===

用户任务：{args.task}

请按照上述任务要求，在手机执行相应操作。
"""

    print("\n🤖 开始执行任务...")
    print(f"   任务：{args.task[:50]}..." if len(args.task) > 50 else f"   任务：{args.task}")
    print(f"   输出文件：{output_file}")
    print(f"   嵌入截图：{'是' if args.embed_screenshot else '否'}")
    print("-" * 50)

    agent = PhoneAgent(model_config=model_cfg, agent_config=agent_cfg)

    try:
        result = agent.run(full_task)

        if args.embed_screenshot:
            print(f"\n📸 嵌入截图...")
            save_screenshot_and_update_excel(
                excel_path=args.file,
                output_path=output_file,
                answer=result,
                success=True,
                embed_screenshot=True
            )
        else:
            print(f"\n💾 保存结果...")
            save_result_to_excel(
                excel_path=args.file,
                output_path=output_file,
                answer=result,
                success=True
            )

        print("\n" + "=" * 50)
        print("✅ 任务执行完成")
        print("=" * 50)
        print(f"输出文件：{output_file}")
        if args.embed_screenshot:
            print(f"✅ 截图已嵌入到 Excel 文档中")
        print("=" * 50)

    except KeyboardInterrupt:
        print("\n\n⚠️  用户中断")
    except Exception as e:
        print(f"\n❌ 任务失败：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def run_batch_mode(args, model_cfg, agent_cfg, output_file):
    """批量执行模式"""
    print("\n" + "=" * 60)
    print("📊 批量执行模式 - 逐条处理 Excel 中的问题")
    print("=" * 60)
    print(f"   输入文件：{args.file}")
    print(f"   输出文件：{output_file}")
    print(f"   任务模板：{args.task[:50]}..." if len(args.task) > 50 else f"   任务模板：{args.task}")
    print(f"   嵌入截图：{'是' if args.embed_screenshot else '否'}")
    print(f"   对比答案：{'是' if args.compare_answer else '否'}")
    if args.max_questions > 0:
        print(f"   最大问题数：{args.max_questions}")
    print("=" * 60)

    try:
        results = process_excel_questions(
            excel_path=args.file,
            task_template=args.task,
            output_path=output_file,
            model_cfg=model_cfg,
            agent_cfg=agent_cfg,
            embed_screenshot=args.embed_screenshot,
            compare_answer=args.compare_answer,
            column=args.column
        )

        success_count = sum(1 for r in results if r.get('success', False))
        failed_count = len(results) - success_count

        print("\n" + "=" * 60)
        print("✅ 批量执行完成")
        print("=" * 60)
        print(f"总问题数：{len(results)}")
        print(f"成功：{success_count}")
        print(f"失败：{failed_count}")
        print(f"成功率：{success_count/len(results)*100:.1f}%" if results else "N/A")
        print(f"输出文件：{output_file}")
        print("=" * 60)

    except KeyboardInterrupt:
        print("\n\n⚠️  用户中断")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 批量执行失败：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
