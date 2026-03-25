"""文件处理工具集 - Excel 读取和写入。

注意：此模块提供基本的 Excel 操作功能。
如需批量执行功能，请使用 skills/excel_tools Skill。
"""

import logging
from pathlib import Path
from typing import Any, Optional

from phone_agent.adb.screenshot import Screenshot
from phone_agent.actions.result import ActionResult

logger = logging.getLogger(__name__)

# Try to import pandas
try:
    import pandas as pd
    from openpyxl import load_workbook
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not installed. Excel support will be limited.")


def handle_read_excel(
    action: dict[str, Any],
    screenshot: Screenshot,
    device_id: Optional[str] = None,
    model_config: Any = None,
    agent_config: Any = None,
) -> ActionResult:
    """
    处理读取 Excel 文件动作。

    Args:
        action: 动作字典，包含 file="文件路径", column="列名" (可选)
        screenshot: Screenshot 对象
        device_id: 设备 ID
        model_config: 模型配置
        agent_config: Agent 配置

    Returns:
        ActionResult: 执行结果
    """
    if not PANDAS_AVAILABLE:
        return ActionResult(
            False, False,
            "Excel 工具不可用：需要安装 pandas 和 openpyxl。请运行：pip install pandas openpyxl"
        )

    file_path = action.get("file")
    column = action.get("column", None)

    if not file_path:
        return ActionResult(False, False, "ReadExcel 需要 file 参数")

    path = Path(file_path)
    if not path.exists():
        return ActionResult(False, False, f"文件不存在：{file_path}")

    if path.suffix.lower() not in ['.xlsx', '.xls']:
        return ActionResult(False, False, f"只支持 Excel 文件 (.xlsx, .xls): {file_path}")

    try:
        df = pd.read_excel(path)
        columns = list(df.columns)
        row_count = len(df)

        # 确定读取的列
        if column:
            if column not in columns:
                return ActionResult(
                    False, False,
                    f"列 '{column}' 不存在。可用列：{columns}"
                )
            items = df[column].dropna().tolist()
            preview_items = items[:3]
            preview = ", ".join([str(item) for item in preview_items])
            content_summary = f"列 '{column}' 共 {len(items)} 条数据"
        else:
            # 自动查找"问题"列
            question_col = None
            for col in columns:
                if '问题' in col.lower() or 'question' in col.lower():
                    question_col = col
                    break

            if question_col:
                items = df[question_col].dropna().tolist()
                preview_items = items[:3]
                preview = ", ".join([str(item) for item in preview_items])
                content_summary = f"找到'{question_col}'列，共 {len(items)} 条数据"
                column = question_col
            else:
                preview = f"列：{columns}"
                content_summary = f"共 {row_count} 行，{len(columns)} 列"

        result_message = (
            f"Excel 文件读取成功：{file_path}\n"
            f"列：{columns}\n"
            f"行数：{row_count}\n"
            f"摘要：{content_summary}\n"
            f"前 3 条预览：{preview}\n"
            f"\n请决定如何处理：\n"
            f"- 批量执行：do(action=\"Execute_Excel_Batch\", file=\"{file_path}\", task=\"任务模板\", column=\"{column or '问题'}\")\n"
            f"- 逐个处理：根据内容调用其他工具"
        )

        logger.info(f"Excel 读取成功：{file_path}, {row_count}行")
        return ActionResult(True, False, message=result_message)

    except Exception as e:
        logger.error(f"Excel 读取失败：{e}", exc_info=True)
        return ActionResult(False, False, f"Excel 读取失败：{e}")


def handle_get_excel_question(
    action: dict[str, Any],
    screenshot: Screenshot,
    device_id: Optional[str] = None,
    model_config: Any = None,
    agent_config: Any = None,
) -> ActionResult:
    """
    获取 Excel 中下一道待处理的问题。

    命令格式：do(action="GetExcelQuestion", file="文件路径", row="行号" (可选))

    - 自动查找状态为"待处理"或"空"的行
    - 返回问题内容和行号，供后续处理
    """
    if not PANDAS_AVAILABLE:
        return ActionResult(False, False, "Excel 工具不可用")

    file_path = action.get("file")
    target_row = action.get("row", None)  # 可选，指定行号

    if not file_path:
        return ActionResult(False, False, "GetExcelQuestion 需要 file 参数")

    path = Path(file_path)
    if not path.exists():
        return ActionResult(False, False, f"文件不存在：{file_path}")

    try:
        df = pd.read_excel(path)

        # 查找问题列
        question_col = None
        for col in df.columns:
            if '问题' in str(col).lower() or 'question' in str(col).lower():
                question_col = col
                break

        if not question_col:
            return ActionResult(False, False, "未找到'问题'列")

        # 查找状态列
        status_col = None
        for col in df.columns:
            if '状态' in str(col).lower() or 'status' in str(col).lower():
                status_col = col
                break

        # 确定要读取的行
        if target_row is not None:
            # 指定行号
            row_idx = int(target_row) - 2  # Excel 行号转索引（减 2：行号从 1 开始，索引从 0 开始，且第 1 行是表头）
            if row_idx < 0 or row_idx >= len(df):
                return ActionResult(False, False, f"行号 {target_row} 超出范围")

            question = str(df.iloc[row_idx][question_col])
            return ActionResult(
                True, False,
                message=f"第{target_row}行的问题：{question}\n\n处理完后请使用 do(action=\"WriteExcelAnswer\", file=\"{file_path}\", row={target_row}, answer=\"答案内容\") 保存答案"
            )
        else:
            # 自动查找下一道待处理的题目
            for idx, row in df.iterrows():
                question = str(row[question_col])

                # 跳过空行
                if not question or question == 'nan':
                    continue

                # 如果有状态列，检查是否已处理
                if status_col:
                    status = str(row[status_col])
                    if status not in ['待处理', '', 'nan', '失败']:
                        continue

                row_num = idx + 2  # 转换为 Excel 行号
                return ActionResult(
                    True, False,
                    message=f"📋 待处理问题 (第{row_num}行):\n{question}\n\n处理完后请使用:\ndo(action=\"WriteExcelAnswer\", file=\"{file_path}\", row={row_num}, answer=\"答案内容\")"
                )

            return ActionResult(True, False, message="✅ 所有问题都已处理完成!")

    except Exception as e:
        logger.error(f"获取问题失败：{e}", exc_info=True)
        return ActionResult(False, False, f"获取问题失败：{e}")


def handle_write_excel_answer(
    action: dict[str, Any],
    screenshot: Screenshot,
    device_id: Optional[str] = None,
    model_config: Any = None,
    agent_config: Any = None,
) -> ActionResult:
    """
    将答案写入 Excel 指定行。

    命令格式：do(action="WriteExcelAnswer", file="文件路径", row="行号", answer="答案内容")
    """
    if not PANDAS_AVAILABLE:
        return ActionResult(False, False, "Excel 工具不可用")

    file_path = action.get("file")
    target_row = action.get("row")
    answer = action.get("answer", "")

    if not file_path:
        return ActionResult(False, False, "WriteExcelAnswer 需要 file 参数")
    if not target_row:
        return ActionResult(False, False, "WriteExcelAnswer 需要 row 参数")
    if not answer:
        return ActionResult(False, False, "WriteExcelAnswer 需要 answer 参数")

    path = Path(file_path)
    if not path.exists():
        return ActionResult(False, False, f"文件不存在：{file_path}")

    try:
        # 使用 openpyxl 读取和修改
        wb = load_workbook(path)
        ws = wb.active

        row_idx = int(target_row)  # Excel 行号（从 1 开始）

        # 查找列索引
        answer_col = 1  # 默认写入第 1 列
        status_col = None

        for col_idx, cell in enumerate(ws[1], 1):  # 遍历第 1 行（表头）
            cell_value = str(cell.value).lower() if cell.value else ''
            if '答案' in cell_value and '标准' not in cell_value:
                answer_col = col_idx
            if '状态' in cell_value or 'status' in cell_value:
                status_col = col_idx

        # 写入答案
        ws.cell(row=row_idx, column=answer_col, value=answer)

        # 更新状态
        if status_col:
            ws.cell(row=row_idx, column=status_col, value='成功')

        # 保存文件
        wb.save(path)
        wb.close()

        logger.info(f"答案已写入 {file_path} 第{target_row}行")

        return ActionResult(
            True, False,
            message=f"✅ 答案已保存到 {file_path} 第{target_row}行\n\n继续处理下一题：do(action=\"GetExcelQuestion\", file=\"{file_path}\")"
        )

    except Exception as e:
        logger.error(f"写入答案失败：{e}", exc_info=True)
        return ActionResult(False, False, f"写入答案失败：{e}")
