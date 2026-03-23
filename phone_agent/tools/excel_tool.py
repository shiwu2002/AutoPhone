"""Excel 批量处理工具 - 供智能体调用。"""

import os
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

# 避免循环导入，使用相对导入
from phone_agent.agent import PhoneAgent, AgentConfig
from phone_agent.model import ModelConfig
from phone_agent.device_factory import get_device_factory
from phone_agent.utils.logger import setup_logger

logger = setup_logger(__name__)

# Try to import pandas for Excel support
try:
    import pandas as pd
    from openpyxl import load_workbook
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not installed. Excel support will be limited.")


@dataclass
class ExcelBatchResult:
    """Excel 批量执行的结果。"""
    total: int
    success: int
    failed: int
    results: list[dict[str, Any]]
    output_path: str
    success_rate: float

    def to_summary(self) -> str:
        """生成结果摘要。"""
        return (
            f"Excel 批量执行完成：共 {self.total} 条，"
            f"成功 {self.success} 条，失败 {self.failed} 条，"
            f"成功率 {self.success_rate:.1f}%"
        )


class ExcelTool:
    """
    Excel 批量处理工具 - 供智能体调用。

    智能体可以通过调用此工具来批量处理 Excel 文件中的任务。

    Example:
        >>> tool = ExcelTool(model_config, agent_config)
        >>> result = tool.execute_batch(
        ...     file_path="questions.xlsx",
        ...     task_template="请回答：{content}",
        ...     question_column="问题"
        ... )
        >>> print(result.to_summary())
    """

    def __init__(
        self,
        model_config: Optional[ModelConfig] = None,
        agent_config: Optional[AgentConfig] = None,
    ):
        """
        初始化 Excel 工具。

        Args:
            model_config: 模型配置
            agent_config: Agent 配置
        """
        self.model_config = model_config or ModelConfig()
        self.agent_config = agent_config or AgentConfig()

    def execute_batch(
        self,
        file_path: str,
        task_template: str,
        question_column: str = "问题",
        output_path: Optional[str] = None,
        embed_screenshot: bool = False,
        compare_answer: bool = False,
        answer_column: str = "答案",
        standard_answer_column: str = "标准答案",
        status_column: str = "状态",
        similarity_column: str = "相似度",
        max_questions: int = 0,
    ) -> ExcelBatchResult:
        """
        批量执行 Excel 中的任务。

        Args:
            file_path: Excel 文件路径
            task_template: 任务模板，可使用 {content} 占位符
            question_column: 问题所在的列名，默认"问题"
            output_path: 输出文件路径，默认为输入文件同名_results.xlsx
            embed_screenshot: 是否嵌入截图，默认 False
            compare_answer: 是否对比标准答案，默认 False
            answer_column: 答案列名，默认"答案"
            standard_answer_column: 标准答案列名，默认"标准答案"
            status_column: 状态列名，默认"状态"
            similarity_column: 相似度列名，默认"相似度"
            max_questions: 最大问题数，0 表示全部

        Returns:
            ExcelBatchResult: 执行结果

        Raises:
            ImportError: pandas 未安装
            FileNotFoundError: 文件不存在
            ValueError: 列名不存在或文件格式错误
        """
        if not PANDAS_AVAILABLE:
            raise ImportError(
                "需要安装 pandas 和 openpyxl: pip install pandas openpyxl"
            )

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在：{file_path}")

        if path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"只支持 Excel 文件 (.xlsx, .xls): {file_path}")

        # 读取 Excel 文件
        df = pd.read_excel(path)

        # 验证列名
        if question_column not in df.columns:
            raise ValueError(
                f"列 '{question_column}' 不存在。可用列：{list(df.columns)}"
            )

        # 获取问题列表
        questions = df[question_column].dropna().astype(str).tolist()
        questions = [q.strip() for q in questions if q.strip() and q != 'nan']

        if not questions:
            raise ValueError("没有找到任何问题")

        # 获取标准答案（如果需要对比）
        standard_answers = []
        if compare_answer:
            if standard_answer_column in df.columns:
                standard_answers = df[standard_answer_column].dropna().astype(str).tolist()
                standard_answers = [s.strip() for s in standard_answers if s.strip() and s != 'nan']
            else:
                logger.warning(f"未找到'{standard_answer_column}'列，将不进行答案对比")
                compare_answer = False

        # 限制问题数量
        if max_questions > 0:
            questions = questions[:max_questions]
            if standard_answers:
                standard_answers = standard_answers[:max_questions]

        logger.info(f"共找到 {len(questions)} 个问题")

        # 确定输出文件路径
        if output_path is None:
            output_path = str(path.parent / f"{path.stem}_results{path.suffix}")

        # 创建 Agent 实例
        agent = PhoneAgent(
            model_config=self.model_config,
            agent_config=self.agent_config,
        )

        # 执行批量任务
        results = []
        success_count = 0
        failed_count = 0

        for i, question in enumerate(questions, 1):
            logger.info(f"\n{'='*60}")
            logger.info(f"问题 {i}/{len(questions)}: {question[:50]}...")
            logger.info(f"{'='*60}")

            try:
                # 构建任务
                if compare_answer and i-1 < len(standard_answers):
                    standard_answer = standard_answers[i-1]
                    if "{content}" in task_template:
                        full_task = self._build_compare_task(
                            task_template.replace("{content}", question),
                            standard_answer
                        )
                    else:
                        full_task = self._build_compare_task(
                            f"{task_template}\n\n问题：{question}",
                            standard_answer
                        )
                else:
                    if "{content}" in task_template:
                        full_task = task_template.replace("{content}", question)
                    else:
                        full_task = f"{task_template}\n\n问题：{question}"

                # 执行任务
                answer = agent.run(full_task)

                # 提取相似度（如果需要对比）
                similarity = None
                if compare_answer:
                    similarity = self._extract_similarity(answer)

                result = {
                    'question': question,
                    'answer': answer,
                    'standard_answer': standard_answers[i-1] if compare_answer and i-1 < len(standard_answers) else None,
                    'similarity': similarity,
                    'success': True,
                    'screenshot_base64': None,
                    'steps': agent.step_count,
                    'error': None,
                }

                # 获取截图（如果启用）
                if embed_screenshot:
                    try:
                        device_factory = get_device_factory()
                        screenshot = device_factory.get_screenshot(enable_compression=False)
                        result['screenshot_base64'] = screenshot.base64_data
                    except Exception as e:
                        logger.warning(f"获取截图失败：{e}")

                results.append(result)
                success_count += 1
                logger.info(f"✅ 完成：{answer[:50] if answer else '无结果'}...")

                # 重置 Agent 状态
                agent.reset()

            except Exception as e:
                logger.error(f"❌ 执行失败：{e}")
                failed_count += 1
                results.append({
                    'question': question,
                    'answer': '',
                    'standard_answer': None,
                    'similarity': None,
                    'success': False,
                    'screenshot_base64': None,
                    'steps': 0,
                    'error': str(e),
                })

        # 保存结果到 Excel
        logger.info(f"\n📊 保存结果到 {output_path}...")
        self._save_results_to_excel(
            path, output_path, results,
            question_column, answer_column, status_column, similarity_column,
            embed_screenshot
        )

        success_rate = (success_count / len(results) * 100) if results else 0.0

        return ExcelBatchResult(
            total=len(results),
            success=success_count,
            failed=failed_count,
            results=results,
            output_path=output_path,
            success_rate=success_rate,
        )

    def _build_compare_task(self, task: str, standard_answer: str) -> str:
        """构建答案对比任务。"""
        return f"""{task}

【答案对比任务】
标准答案：{standard_answer}

请你对比上面获取的答案和标准答案，计算相似度（0-100 的分数），考虑：
1. 关键信息是否一致
2. 核心要点是否覆盖

请输出 JSON 格式的结果：
{{"答案": "你的答案", "相似度": 85}}
"""

    def _extract_similarity(self, answer: str) -> Optional[int]:
        """从答案中提取相似度分数。"""
        import re
        try:
            json_match = re.search(r'"相似度"\s*:\s*(\d+)', answer)
            if json_match:
                return int(json_match.group(1))
        except Exception as e:
            logger.warning(f"提取相似度失败：{e}")
        return None

    def _save_results_to_excel(
        self,
        input_path: Path,
        output_path: str,
        results: list[dict],
        question_column: str,
        answer_column: str,
        status_column: str,
        similarity_column: str,
        embed_screenshot: bool,
    ) -> None:
        """保存结果到 Excel 文件。"""
        df = pd.read_excel(input_path)

        # 添加或更新列
        if answer_column not in df.columns:
            df[answer_column] = ''
        if status_column not in df.columns:
            df[status_column] = ''
        if similarity_column not in df.columns:
            df[similarity_column] = ''

        # 更新数据
        for i, result in enumerate(results):
            if i < len(df):
                df.loc[i, answer_column] = result.get('answer', '')
                df.loc[i, status_column] = (
                    '成功' if result.get('success', False)
                    else f"失败：{result.get('error', '')}"
                )
                if result.get('similarity') is not None:
                    df.loc[i, similarity_column] = result['similarity']

        # 保存 Excel
        df.to_excel(output_path, index=False, engine='openpyxl')
        logger.info(f"✅ 结果已保存到：{output_path}")

        # 嵌入截图（如果启用）
        if embed_screenshot:
            self._embed_screenshots_to_excel(output_path, results)

    def _embed_screenshots_to_excel(
        self,
        output_path: str,
        results: list[dict],
    ) -> None:
        """将截图嵌入到 Excel 文件。"""
        import base64
        import io
        from PIL import Image as PILImage

        try:
            wb = load_workbook(output_path)
            ws = wb.active

            # 查找截图列
            screenshot_col = None
            headers = {cell.column_letter: cell.value for cell in ws[1]}
            for col_letter, header in headers.items():
                if header == '截图':
                    screenshot_col = col_letter
                    break

            if screenshot_col is None:
                # 添加截图列
                max_col = max(cell.column for cell in ws[1])
                screenshot_col_letter = chr(ord('A') + max_col)
                ws.cell(row=1, column=max_col + 1, value='截图')
                screenshot_col = screenshot_col_letter

            temp_paths = []

            for i, result in enumerate(results):
                screenshot_b64 = result.get('screenshot_base64')
                if screenshot_b64:
                    row = i + 2
                    try:
                        # 解码图片
                        image_data = base64.b64decode(screenshot_b64)
                        img = PILImage.open(io.BytesIO(image_data))

                        # 缩放图片
                        max_width = 300
                        max_height = 400
                        width, height = img.size
                        ratio = min(max_width / width, max_height / height)
                        if ratio < 1:
                            new_width = int(width * ratio * 0.75)
                            new_height = int(height * ratio * 0.75)
                        else:
                            new_width = int(width * 0.75)
                            new_height = int(height * 0.75)

                        # 保存临时文件
                        temp_path = Path(output_path).parent / f"temp_screenshot_{i}.png"
                        img.save(temp_path)
                        temp_paths.append(temp_path)

                        # 嵌入到 Excel
                        from openpyxl.drawing.image import Image
                        img_obj = Image(temp_path)
                        img_obj.width = new_width
                        img_obj.height = new_height
                        anchor = f"{screenshot_col}{row}"
                        img_obj.anchor = anchor
                        ws.add_image(img_obj)
                        ws.row_dimensions[row].height = new_height / 6

                    except Exception as e:
                        logger.warning(f"嵌入截图失败 (行{row}): {e}")

            # 保存 Excel
            wb.save(output_path)

            # 删除临时文件
            for temp_path in temp_paths:
                try:
                    if temp_path.exists():
                        os.remove(temp_path)
                except:
                    pass

            logger.info(f"✅ 截图已嵌入到 Excel")
        except Exception as e:
            logger.warning(f"嵌入截图失败：{e}")


# 便捷函数
def execute_excel_batch(
    file_path: str,
    task_template: str,
    question_column: str = "问题",
    output_path: Optional[str] = None,
    embed_screenshot: bool = False,
    compare_answer: bool = False,
    max_questions: int = 0,
    model_config: Optional[ModelConfig] = None,
    agent_config: Optional[AgentConfig] = None,
) -> ExcelBatchResult:
    """
    便捷函数：批量执行 Excel 中的任务。

    Args:
        file_path: Excel 文件路径
        task_template: 任务模板，可使用 {content} 占位符
        question_column: 问题所在的列名
        output_path: 输出文件路径
        embed_screenshot: 是否嵌入截图
        compare_answer: 是否对比标准答案
        max_questions: 最大问题数，0 表示全部
        model_config: 模型配置
        agent_config: Agent 配置

    Returns:
        ExcelBatchResult: 执行结果
    """
    tool = ExcelTool(model_config=model_config, agent_config=agent_config)
    return tool.execute_batch(
        file_path=file_path,
        task_template=task_template,
        question_column=question_column,
        output_path=output_path,
        embed_screenshot=embed_screenshot,
        compare_answer=compare_answer,
        max_questions=max_questions,
    )
